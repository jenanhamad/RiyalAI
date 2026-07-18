"""Business import/export: bring in past expense reports (CSV/Excel) and
export current data back out (raw data or a summary report).

Flow:
  1. POST /business/import/preview  -> parse_upload() + suggest a column mapping,
     stash the parsed rows on disk keyed by an importId.
  2. POST /business/import/confirm  -> re-load the stash, apply the (possibly
     user-edited) mapping, normalize + insert rows, delete the stash.

Export:
  export_expenses_file() -> raw rows as CSV/XLSX
  export_report_file()   -> a formatted XLSX summary (profit, VAT, categories, leaks)
"""
from __future__ import annotations

import csv
import io
import json
import re
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import business as biz
from database import get_connection

_DATA_ROOT = Path(
    __import__("os").environ.get(
        "DATA_DIR", str(Path(__file__).parent.parent / "data")
    )
)
IMPORT_DIR = _DATA_ROOT / "imports"
IMPORT_DIR.mkdir(parents=True, exist_ok=True)

MAX_IMPORT_ROWS = 5000
IMPORT_SESSION_TTL = timedelta(hours=24)

# Fields the user can map source columns to.
TARGET_FIELDS = [
    "merchant", "amount", "date", "category", "entryType",
    "paymentMethod", "description", "projectTag",
]

REQUIRED_TARGET_FIELDS = ("amount",)

_ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")

_HEADER_HINTS = {
    "merchant": ["merchant", "vendor", "supplier", "payee", "الوصف", "المورد", "الجهة", "العميل", "التاجر", "البيان", "المستفيد"],
    "amount": ["amount", "total", "value", "price", "sar", "المبلغ", "القيمة", "الإجمالي", "ريال", "سعر"],
    "date": ["date", "day", "التاريخ", "تاريخ"],
    "category": ["category", "type of expense", "الفئة", "التصنيف", "النوع", "البند"],
    "entryType": ["entry type", "income/expense", "kind", "نوع الحركة", "دخل/مصروف", "مصروف/إيراد"],
    "paymentMethod": ["payment", "method", "طريقة الدفع", "وسيلة الدفع"],
    "description": ["description", "notes", "memo", "وصف", "ملاحظات", "بيان"],
    "projectTag": ["project", "client", "tag", "مشروع", "عميل", "وسم"],
}

_ENTRY_TYPE_INCOME_HINTS = ("income", "revenue", "sale", "إيراد", "دخل", "بيع", "مبيعات", "تحصيل")
_ENTRY_TYPE_EXPENSE_HINTS = ("expense", "cost", "مصروف", "مصاريف", "صرف", "شراء", "تكلفة")

_DATE_FORMATS = [
    "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m-%d-%Y",
    "%d.%m.%Y", "%Y.%m.%d", "%d/%m/%y", "%d-%m-%y",
]


class ImportError_(ValueError):
    pass


# ---------------------------------------------------------------------------
# Parsing uploaded files
# ---------------------------------------------------------------------------

def parse_upload(filename: str, content: bytes) -> tuple[list[str], list[list[str]]]:
    """Return (columns, rows) with every cell coerced to a plain string."""
    lower = (filename or "").lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _parse_xlsx(content)
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return _parse_csv(content)
    # Best effort: sniff — try CSV first, then xlsx
    try:
        return _parse_csv(content)
    except Exception:
        return _parse_xlsx(content)


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_xlsx(content: bytes) -> tuple[list[str], list[list[str]]]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ImportError_("الملف فاضي")
    columns = [(_cell_to_str(c) or f"عمود {i + 1}") for i, c in enumerate(header)]
    rows = []
    for row in rows_iter:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        vals = [_cell_to_str(c) for c in row]
        if len(vals) < len(columns):
            vals += [""] * (len(columns) - len(vals))
        rows.append(vals[: len(columns)])
        if len(rows) >= MAX_IMPORT_ROWS:
            break
    return columns, rows


def _parse_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = content.decode("utf-8", errors="replace")

    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration:
        raise ImportError_("الملف فاضي")
    columns = [(c.strip() or f"عمود {i + 1}") for i, c in enumerate(header)]
    rows = []
    for row in reader:
        if not row or all((c or "").strip() == "" for c in row):
            continue
        vals = [c.strip() for c in row]
        if len(vals) < len(columns):
            vals += [""] * (len(columns) - len(vals))
        rows.append(vals[: len(columns)])
        if len(rows) >= MAX_IMPORT_ROWS:
            break
    return columns, rows


# ---------------------------------------------------------------------------
# Heuristic column mapping (fallback when AI is unavailable)
# ---------------------------------------------------------------------------

def heuristic_mapping(columns: list[str]) -> dict[str, str | None]:
    mapping: dict[str, str | None] = {f: None for f in TARGET_FIELDS}
    used = set()
    for field, hints in _HEADER_HINTS.items():
        best = None
        for col in columns:
            if col in used:
                continue
            norm = col.strip().lower()
            if any(h in norm for h in hints):
                best = col
                break
        if best:
            mapping[field] = best
            used.add(best)
    return mapping


# ---------------------------------------------------------------------------
# Import session persistence (short-lived stash between preview + confirm)
# ---------------------------------------------------------------------------

def _session_path(user_id: str, import_id: str) -> Path:
    return IMPORT_DIR / user_id / f"{import_id}.json"


def save_import_session(user_id: str, filename: str, columns: list[str], rows: list[list[str]]) -> str:
    _cleanup_expired_sessions(user_id)
    import_id = str(uuid.uuid4())
    path = _session_path(user_id, import_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps({
            "userId": user_id,
            "filename": filename,
            "columns": columns,
            "rows": rows,
            "createdAt": datetime.utcnow().isoformat(),
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    return import_id


def load_import_session(user_id: str, import_id: str) -> dict:
    path = _session_path(user_id, import_id)
    if not path.is_file():
        raise ImportError_("انتهت صلاحية الملف — أعد رفعه")
    return json.loads(path.read_text(encoding="utf-8"))


def delete_import_session(user_id: str, import_id: str) -> None:
    path = _session_path(user_id, import_id)
    if path.is_file():
        path.unlink()


def _cleanup_expired_sessions(user_id: str) -> None:
    folder = IMPORT_DIR / user_id
    if not folder.is_dir():
        return
    cutoff = datetime.utcnow() - IMPORT_SESSION_TTL
    for f in folder.glob("*.json"):
        try:
            if datetime.utcfromtimestamp(f.stat().st_mtime) < cutoff:
                f.unlink()
        except OSError:
            continue


# ---------------------------------------------------------------------------
# Row normalization
# ---------------------------------------------------------------------------

def _parse_amount(raw: str) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip().translate(_ARABIC_DIGITS)
    if not s:
        return None
    negative = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in ("-", "."):
        return None
    try:
        val = float(s)
    except ValueError:
        return None
    return -abs(val) if negative else val


def _parse_date(raw: str) -> str | None:
    if not raw:
        return None
    s = str(raw).strip().translate(_ARABIC_DIGITS)
    if not s:
        return None
    if "T" in s:
        s = s.split("T")[0]
    for fmt in _DATE_FORMATS:
        try:
            dt = datetime.strptime(s, fmt)
            year = dt.year if dt.year > 100 else dt.year + 2000
            return date(year, dt.month, dt.day).isoformat()
        except ValueError:
            continue
    # Excel serial date fallback (days since 1899-12-30)
    try:
        serial = float(s)
        if 1000 < serial < 80000:
            base = date(1899, 12, 30)
            return (base + timedelta(days=int(serial))).isoformat()
    except ValueError:
        pass
    return None


def _parse_entry_type(raw: str, default: str) -> str:
    if not raw:
        return default
    norm = str(raw).strip().lower()
    if any(h in norm for h in _ENTRY_TYPE_INCOME_HINTS):
        return "income"
    if any(h in norm for h in _ENTRY_TYPE_EXPENSE_HINTS):
        return "expense"
    return default


def normalize_rows(
    columns: list[str],
    rows: list[list[str]],
    mapping: dict[str, str | None],
    *,
    default_entry_type: str = "expense",
    default_category: str = "Other",
) -> tuple[list[dict], list[dict]]:
    """Returns (normalized_rows, row_errors)."""
    col_index = {c: i for i, c in enumerate(columns)}
    normalized = []
    errors = []
    today = date.today().isoformat()

    for i, row in enumerate(rows):
        def get(field):
            col = mapping.get(field)
            if not col or col not in col_index:
                return ""
            idx = col_index[col]
            return row[idx] if idx < len(row) else ""

        amount_raw = get("amount")
        amount = _parse_amount(amount_raw)
        if amount is None or amount == 0:
            errors.append({"row": i + 2, "reason": "مبلغ غير صالح", "raw": amount_raw})
            continue

        has_entry_type_column = bool(mapping.get("entryType"))
        entry_type = _parse_entry_type(get("entryType"), default_entry_type)
        if amount < 0 and not has_entry_type_column:
            # No explicit type column — treat negative amounts as expenses (outflow).
            entry_type = "expense"
        amount = abs(amount)

        merchant = (get("merchant") or "").strip()
        description = (get("description") or "").strip()
        if not merchant:
            merchant = description or ("إيراد مستورد" if entry_type == "income" else "مصروف مستورد")

        parsed_date = _parse_date(get("date")) or today
        category_raw = (get("category") or "").strip()
        category = biz.normalize_business_category(category_raw) if category_raw else default_category
        payment_method = (get("paymentMethod") or "").strip() or "تحويل بنكي"
        project_tag = (get("projectTag") or "").strip()[:80]

        normalized.append({
            "merchant": merchant[:100],
            "amount": round(amount, 2),
            "date": parsed_date,
            "category": category,
            "entryType": entry_type,
            "paymentMethod": payment_method[:40],
            "description": description[:300],
            "projectTag": project_tag,
        })

    return normalized, errors


def existing_signatures(user_id: str, mode: str = "business") -> set[tuple]:
    conn = get_connection()
    rows = conn.execute(
        "SELECT merchant, amount, date FROM expenses WHERE user_id = ? AND mode = ?",
        (user_id, mode),
    ).fetchall()
    conn.close()
    return {
        (str(r["merchant"]).strip().lower(), round(float(r["amount"]), 2), str(r["date"])[:10])
        for r in rows
    }


def row_signature(row: dict) -> tuple:
    return (row["merchant"].strip().lower(), round(float(row["amount"]), 2), row["date"][:10])


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

CSV_HEADERS = [
    ("date", "التاريخ"), ("entryType", "النوع"), ("merchant", "الوصف / المستفيد"),
    ("amount", "المبلغ"), ("category", "الفئة"), ("paymentMethod", "طريقة الدفع"),
    ("projectTag", "المشروع / العميل"), ("description", "ملاحظات"), ("hasReceipt", "إيصال؟"),
]


def _fetch_export_rows(user_id: str, mode: str, since: str | None) -> list[dict]:
    conn = get_connection()
    if since:
        rows = conn.execute(
            "SELECT * FROM expenses WHERE user_id = ? AND mode = ? AND date >= ? ORDER BY date DESC",
            (user_id, mode, since),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM expenses WHERE user_id = ? AND mode = ? ORDER BY date DESC",
            (user_id, mode),
        ).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        out.append({
            "date": str(d.get("date") or "")[:10],
            "entryType": "إيراد" if (d.get("entry_type") or "expense") == "income" else "مصروف",
            "merchant": d.get("merchant") or "",
            "amount": float(d.get("amount") or 0),
            "category": d.get("category") or "",
            "paymentMethod": d.get("payment_method") or "",
            "projectTag": d.get("project_tag") or "",
            "description": d.get("description") or d.get("notes") or "",
            "hasReceipt": "نعم" if d.get("has_receipt") else "لا",
        })
    return out


def export_expenses_file(user_id: str, *, mode: str = "business", fmt: str = "xlsx", days: int | None = None):
    since = (date.today() - timedelta(days=days)).isoformat() if days else None
    rows = _fetch_export_rows(user_id, mode, since)
    stamp = date.today().isoformat()

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([label for _, label in CSV_HEADERS])
        for r in rows:
            writer.writerow([r[key] for key, _ in CSV_HEADERS])
        data = ("\ufeff" + buf.getvalue()).encode("utf-8")
        return data, f"riyalai_export_{stamp}.csv", "text/csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "الحركات"
    ws.sheet_view.rightToLeft = True
    header_fill = PatternFill("solid", fgColor="1F6F5C")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append([label for _, label in CSV_HEADERS])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r[key] for key, _ in CSV_HEADERS])
    for i, _ in enumerate(CSV_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"riyalai_export_{stamp}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def export_report_file(user_id: str, *, days: int = 90):
    """A printable XLSX summary report: profit, VAT, categories, leaks, entries."""
    profit = biz.profit_snapshot(user_id, days=days)
    vat = biz.vat_summary(user_id, days=days)
    health = biz.health_score(user_id)
    leaks = biz.rule_based_leaks(user_id)
    rows = _fetch_export_rows(user_id, "business", (date.today() - timedelta(days=days)).isoformat())
    stamp = date.today().isoformat()

    wb = Workbook()
    ws = wb.active
    ws.title = "الملخص"
    ws.sheet_view.rightToLeft = True
    title_font = Font(bold=True, size=14, color="1F6F5C")
    label_font = Font(bold=True)

    def add_row(label, value=""):
        ws.append([label, value])

    add_row(f"تقرير مشروع — آخر {days} يوم", "")
    ws["A1"].font = title_font
    add_row("تاريخ التصدير", stamp)
    add_row("")
    add_row("الإيرادات", profit["income"])
    add_row("المصروفات", profit["expenses"])
    add_row("الربح التقريبي", profit["profit"])
    add_row("هامش الربح %", profit["marginPercent"])
    add_row("صحة المشروع (0-100)", health["score"])
    add_row("VAT قابل للاسترداد (تقديري)", vat["vatRecoverableEstimate"])
    add_row("مصروفات بدون إيصال", vat["missingReceiptCount"])
    add_row("")
    add_row("تفصيل المصروفات حسب الفئة")
    for cat, amt in profit["categoryBreakdown"].items():
        add_row(f"  {cat}", round(amt, 2))
    if profit["projectBreakdown"]:
        add_row("")
        add_row("حسب المشروع / العميل")
        for tag, amt in profit["projectBreakdown"].items():
            add_row(f"  {tag}", round(amt, 2))
    if leaks:
        add_row("")
        add_row("نقاط هدر محتملة")
        for leak in leaks:
            add_row(f"  {leak['title']}", leak.get("amount", ""))
            ws.append(["  " + leak.get("suggestion", ""), ""])

    for row in ws.iter_rows(min_row=4):
        if row[0].value and str(row[0].value).strip() and not str(row[0].value).startswith("  "):
            row[0].font = label_font
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20

    ws2 = wb.create_sheet("الحركات التفصيلية")
    ws2.sheet_view.rightToLeft = True
    ws2.append([label for _, label in CSV_HEADERS])
    header_fill = PatternFill("solid", fgColor="1F6F5C")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in rows:
        ws2.append([r[key] for key, _ in CSV_HEADERS])
    for i, _ in enumerate(CSV_HEADERS, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"riyalai_report_{stamp}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
